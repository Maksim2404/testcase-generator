import os, csv, pathlib, re, sys
from collections import Counter

# 3rd-party (CI installs them)
try:
    import yaml
    from jinja2 import Template
except Exception:
    print("Missing deps. Locally: pip install pyyaml jinja2")
    raise

ROOT = pathlib.Path(__file__).resolve().parents[1]
CASES_GLOB = "testcase-generator/apps/**/**/*.md"
CI_SERVER_URL = os.getenv("CI_SERVER_URL", "").rstrip("/")

WARNINGS = []

def write_stats(rows, out_path_json, out_path_html):
    by_app = Counter(r.get("app","") for r in rows)
    by_suite = Counter(r.get("suite","") for r in rows)
    by_prio = Counter(r.get("priority","") for r in rows)
    data = {"by_app": by_app, "by_suite": by_suite, "by_priority": by_prio, "total": len(rows)}
    import json
    out_path_json.parent.mkdir(parents=True, exist_ok=True)
    out_path_json.write_text(json.dumps({k: dict(v) if hasattr(v,'items') else v for k,v in data.items()}, indent=2), encoding="utf-8")
    # barebones HTML:
    html = f"""<h3>Total: {data['total']}</h3>
<ul>
  <li>Apps: {dict(by_app)}</li>
  <li>Suites: {dict(by_suite)}</li>
  <li>Priorities: {dict(by_prio)}</li>
</ul>"""
    out_path_html.write_text(html, encoding="utf-8")

# BOM/whitespace + CRLF tolerant front-matter
FRONT_MATTER_RE = re.compile(
    r'^\ufeff?\s*---\s*\r?\n(.*?)\r?\n---\s*(?:\r?\n|$)',
    re.DOTALL | re.MULTILINE
)
ISSUE_REF_RE = re.compile(r'^([^#\s]+)#(\d+)$')  # group/project#123

def parse_case(md_path: pathlib.Path):
    """Parse one Markdown file with optional YAML front-matter and H1 title."""
    text = md_path.read_text(encoding="utf-8")
    meta = {}
    title = md_path.stem

    mfm = FRONT_MATTER_RE.search(text)
    if not mfm:
        WARNINGS.append(f"No front matter: {md_path}")
    else:
        try:
            meta = yaml.safe_load(mfm.group(1)) or {}
        except Exception as e:
            WARNINGS.append(f"YAML parse error in {md_path}: {e}")
            meta = {}

    # H1 title (first '# ')
    mh1 = re.search(r'^\s*#\s+(.+)$', text, re.MULTILINE)
    if mh1:
        title = mh1.group(1).strip()

    def _get(d, k, default=""):
        v = d.get(k, default)
        return v if v is not None else default

    def _norm_refs(lst):
        out = []
        for s in (lst or []):
            s = str(s).strip()
            m = ISSUE_REF_RE.match(s)
            if m and CI_SERVER_URL:
                proj, iid = m.group(1), m.group(2)
                url = f"{CI_SERVER_URL}/{proj}/-/issues/{iid}"
                out.append({"text": s, "url": url})
            else:
                out.append({"text": s, "url": ""})
        return out

    return {
        "id": _get(meta,"id",md_path.stem),
        "title": title,
        "app": _get(meta,"app"),
        "area": _get(meta,"area"),
        "suite": _get(meta,"suite"),
        "type": _get(meta,"type"),
        "priority": _get(meta,"priority"),
        "status": _get(meta,"status"),
        "owner": _get(meta,"owner"),
        "stories_list": _norm_refs(_get(meta,"story_refs",[])),
        "bugs_list": _norm_refs(_get(meta,"bug_refs",[])),
        "stories": ", ".join(_get(meta,"story_refs",[]) or []),
        "bugs": ", ".join(_get(meta,"bug_refs",[]) or []),
        "automation": (_get(meta,"automation",{}) or {}).get("status",""),
        "path": str(md_path.relative_to(ROOT)),
    }

def write_csv(rows, out_path):
    out_path.parent.mkdir(parents=True, exist_ok=True)

    # Internal keys (from rows) -> Pretty headers (for CSV)
    cols = [
        ("id","ID"), ("title","Title"), ("app","App"), ("area","Area"),
        ("suite","Suite"), ("type","Type"), ("priority","Priority"),
        ("status","Status"), ("owner","Owner"),
        ("stories","Stories"), ("bugs","Bugs"),
        ("automation","Automation"), ("path","Path"),
    ]

    # sanitize rows and map to pretty headers
    cleaned = []
    for r in rows:
        row = {}
        for key, header in cols:
            row[header] = r.get(key, "")
        cleaned.append(row)

    import io, csv
    # UTF-8 with BOM for Excel
    with out_path.open("w", newline="", encoding="utf-8-sig") as fp:
        if cleaned:
            w = csv.DictWriter(
                fp,
                fieldnames=[h for _, h in cols],
                quoting=csv.QUOTE_ALL,
                lineterminator="\r\n",
            )
            w.writeheader()
            w.writerows(cleaned)
        else:
            fp.write("")

def write_xlsx(rows, out_path):
    from openpyxl import Workbook
    from openpyxl.styles import Font
    from openpyxl.worksheet.dimensions import ColumnDimension

    out_path.parent.mkdir(parents=True, exist_ok=True)

    cols = [
        ("id","ID"), ("title","Title"), ("app","App"), ("area","Area"),
        ("suite","Suite"), ("type","Type"), ("priority","Priority"),
        ("status","Status"), ("owner","Owner"),
        ("stories","Stories"), ("bugs","Bugs"),
        ("automation","Automation"), ("path","Path"),
    ]

    wb = Workbook()
    ws = wb.active
    ws.title = "Traceability"

    headers = [h for _, h in cols]
    ws.append(headers)

    # bold header
    for cell in ws[1]:
        cell.font = Font(bold=True)

    # rows
    for r in rows:
        ws.append([r.get(k, "") for k, _ in cols])

    # freeze header and enable filter
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    # simple column widths
    widths = {
        "A": 14, "B": 42, "C": 12, "D": 26, "E": 12, "F": 14,
        "G": 10, "H": 12, "I": 12, "J": 22, "K": 28, "L": 28, "M": 14, "N": 36
    }
    for col, w in widths.items():
        ws.column_dimensions[col] = ColumnDimension(ws, min=ws[col+'1'].column, max=ws[col+'1'].column, width=w)

    wb.save(out_path)

def write_warnings(out_path):
    out_path.parent.mkdir(parents=True, exist_ok=True)
    with out_path.open("w", encoding="utf-8") as f:
        if WARNINGS:
            f.write("\n".join(WARNINGS))
        else:
            f.write("No warnings.\n")

def write_html(rows, out_path):
    out_path.parent.mkdir(parents=True, exist_ok=True)
    heads = [
        ("id","ID"), ("title","Title"), ("app","App"), ("area","Area"),
        ("suite","Suite"), ("type","Type"), ("priority","Priority"),
        ("status","Status"), ("owner","Owner"),
        ("stories_list","Stories"), ("bugs_list","Bugs"),
        ("automation","Automation"), ("path","Path"),
    ]
    tpl = Template(r"""
<!doctype html><meta charset="utf-8">
<title>Traceability</title>
<style>
:root { --chip:#eef; --chip2:#efe; --hdr:#f6f7fb; }
body{font-family:system-ui,-apple-system,Segoe UI,Roboto,Helvetica,Arial,sans-serif;margin:16px}
h1{margin:0 0 12px 0;font-size:20px}
.controls{display:flex;gap:8px;flex-wrap:wrap;margin:8px 0 12px}
.controls select, .controls input{padding:6px 8px;font-size:14px}
table{border-collapse:collapse;width:100%}
th,td{border:1px solid #e5e7eb;padding:8px 10px;font-size:14px;vertical-align:top}
th{background:var(--hdr); position:sticky; top:0; z-index:1; cursor:pointer; user-select:none}
tbody tr:nth-child(odd){background:#fafafa}
.badge{display:inline-block; padding:2px 8px; border-radius:999px; font-size:12px}
.suite-Smoke{background:var(--chip)}
.suite-Regression{background:var(--chip2)}
.priority-P0{background:#ffe0e0}
.priority-P1{background:#fff2cc}
.priority-P2{background:#e6f7ff}
.priority-P3{background:#eef}
.status-Ready{background:#e8ffe8}
.status-Draft{background:#fde68a}
.cell a{color:#2563eb; text-decoration:none}
.count{opacity:.7;margin-left:6px}
.small{font-size:12px; opacity:.7}
</style>

<h1>Traceability <span class="count">({{rows|length}} cases)</span></h1>

<div class="controls">
  <input id="q" placeholder="Search…" />
  <select id="fApp"><option value="">All Apps</option></select>
  <select id="fSuite"><option value="">All Suites</option></select>
  <select id="fPrio"><option value="">All Priorities</option></select>
  <button id="clear">Clear</button>
</div>

<table id="t" data-sortcol="0" data-sortdir="asc">
  <thead><tr>
    {% for key,label in heads %}<th data-key="{{key}}">{{label}}</th>{% endfor %}
  </tr></thead>
  <tbody>
    {% for r in rows %}
    <tr data-app="{{r['app']}}" data-suite="{{r['suite']}}" data-prio="{{r['priority']}}">
      <td class="cell">{{r["id"]}}</td>
      <td class="cell">{{r["title"]}}</td>
      <td class="cell">{{r["app"]}}</td>
      <td class="cell">{{r["area"]}}</td>
      <td class="cell"><span class="badge suite-{{r["suite"]|e}}">{{r["suite"]}}</span></td>
      <td class="cell">{{r["type"]}}</td>
      <td class="cell"><span class="badge priority-{{r["priority"]|e}}">{{r["priority"]}}</span></td>
      <td class="cell"><span class="badge status-{{r["status"]|e}}">{{r["status"]}}</span></td>
      <td class="cell">{{r["owner"]}}</td>
      <td class="cell">
        {% if r["stories_list"] %}
          {% for s in r["stories_list"] %}
            {% if s.url %}<a href="{{s.url}}">{{s.text}}</a>{% else %}{{s.text}}{% endif %}{% if not loop.last %}, {% endif %}
          {% endfor %}
        {% endif %}
      </td>
      <td class="cell">
        {% if r["bugs_list"] %}
          {% for b in r["bugs_list"] %}
            {% if b.url %}<a href="{{b.url}}">{{b.text}}</a>{% else %}{{b.text}}{% endif %}{% if not loop.last %}, {% endif %}
          {% endfor %}
        {% endif %}
      </td>
      <td class="cell">{{r["automation"]}}</td>
      <td class="cell small">{{r["path"]}}</td>
    </tr>
    {% endfor %}
  </tbody>
</table>

<script>
const q = document.getElementById('q');
const t = document.getElementById('t');
const fApp = document.getElementById('fApp');
const fSuite = document.getElementById('fSuite');
const fPrio = document.getElementById('fPrio');
const clearBtn = document.getElementById('clear');

function buildFilters() {
  const vals = {app:new Set(), suite:new Set(), prio:new Set()};
  for (const tr of t.tBodies[0].rows) {
    vals.app.add(tr.dataset.app || '');
    vals.suite.add(tr.dataset.suite || '');
    vals.prio.add(tr.dataset.prio || '');
  }
  for (const v of [...vals.app].filter(Boolean).sort()) fApp.add(new Option(v,v));
  for (const v of [...vals.suite].filter(Boolean).sort()) fSuite.add(new Option(v,v));
  for (const v of [...vals.prio].filter(Boolean).sort()) fPrio.add(new Option(v,v));
}
buildFilters();

function applyFilters() {
  const term = (q.value||'').toLowerCase();
  const a = fApp.value, s = fSuite.value, p = fPrio.value;
  for (const tr of t.tBodies[0].rows) {
    const hit = (!a || tr.dataset.app===a)
      && (!s || tr.dataset.suite===s)
      && (!p || tr.dataset.prio===p)
      && ([...tr.cells].some(td => td.textContent.toLowerCase().includes(term)));
    tr.style.display = hit ? '' : 'none';
  }
}
[q,fApp,fSuite,fPrio].forEach(el => el.addEventListener('input', applyFilters));
clearBtn.addEventListener('click', ()=>{ q.value=''; fApp.value=''; fSuite.value=''; fPrio.value=''; applyFilters(); });

document.querySelectorAll('th').forEach((th,idx)=>{
  th.addEventListener('click', ()=>{
    const dir = (t.dataset.sortdir==='asc' && t.dataset.sortcol==idx) ? 'desc':'asc';
    t.dataset.sortcol = idx; t.dataset.sortdir = dir;
    const rows = [...t.tBodies[0].rows];
    rows.sort((a,b)=>{
      const av=a.cells[idx].textContent.trim().toLowerCase();
      const bv=b.cells[idx].textContent.trim().toLowerCase();
      if (av<bv) return dir==='asc'?-1:1;
      if (av>bv) return dir==='asc'?1:-1;
      return 0;
    });
    for (const r of rows) t.tBodies[0].appendChild(r);
  });
});
</script>
""")
    out_path.write_text(tpl.render(rows=rows, heads=heads), encoding="utf-8")

def main():
    files = list(ROOT.glob(CASES_GLOB))
    rows = [parse_case(p) for p in files]
    rows.sort(key=lambda r: (r.get("app",""), r.get("id","")))

    out_dir = ROOT/"traceability"
    write_csv(rows, out_dir/"traceability.csv")
    # XLSX optional
    try:
        write_xlsx(rows, out_dir/"traceability.xlsx")
    except Exception as e:
        WARNINGS.append(f"Excel export skipped: {e}")
    write_html(rows, out_dir/"index.html")
    write_warnings(out_dir/"warnings.txt")
    write_stats(rows, ROOT/"traceability/stats.json", ROOT/"traceability/stats.html")

    print(f"Built matrix with {len(rows)} cases from {len(files)} files")
    if WARNINGS and os.getenv("STRICT_YAML") == "1":
        for w in WARNINGS:
            print("[WARN]", w)
        sys.exit(2)

if __name__ == "__main__":
    try:
        main()
    except Exception as e:
        print("ERROR building traceability:", e)
        sys.exit(1)